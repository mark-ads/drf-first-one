import logging
from typing import cast

from django.core.files.uploadedfile import InMemoryUploadedFile
from django.http import HttpRequest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from first_one.first_app.models import Event, EventPlace
from first_one.first_app.serializers import EventPlaceSerializer, EventSerializer

logger = logging.getLogger("first_app")


class EventImportService:
    def __init__(self, file: InMemoryUploadedFile, request: HttpRequest):
        self.sheet = self._get_sheet(file)
        self.request = request
        self.errors = []
        self.created = 0

    def _get_sheet(self, file: InMemoryUploadedFile) -> Worksheet:
        self.workbook = load_workbook(file, read_only=True)
        return cast(Worksheet, self.workbook.active)  # указываем тип для LSP

    def run(self):
        logger.debug("Старт импорта мероприятий")
        for row_number, row in enumerate(
            self.sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if len(row) < 9:
                logger.debug(f"Ошибка импорта в строке {row_number}")
                self.errors.append({"row_number": row_number})
                continue

            (
                event_name,
                description,
                publish_date,
                start_date,
                end_date,
                place_name,
                latitude,
                longitude,
                rating,
            ) = row

            # Проверка на строку, чтобы корректно сделать запрос в БД
            if not isinstance(place_name, str):
                self.errors.append(
                    {"row_number": row_number, "errors": "Название места пустое"}
                )
                continue
            place_name = place_name.strip()
            place = EventPlace.objects.filter(name__iexact=place_name).first()

            if place is None:
                place_data = {
                    "name": place_name,
                    "latitude": latitude,
                    "longitude": longitude,
                }

                serializer = EventPlaceSerializer(data=place_data)

                if serializer.is_valid():
                    logger.info(f"Создано новое место из импорта {place_name}")
                    place = serializer.save()

                else:
                    logger.debug(f"Ошибка создания места из импорта {place_name}")
                    self.errors.append(
                        {"row_number": row_number, "errors": serializer.errors}
                    )
                    continue

            event_data = {
                "name": event_name,
                "description": description,
                "publish_date": publish_date,
                "start_date": start_date,
                "end_date": end_date,
                "place": place.id,  # type: ignore
                "rating": rating,
                "status": Event.StatusChoices.DRAFT,
            }

            serializer = EventSerializer(
                data=event_data, context={"request": self.request}
            )

            if serializer.is_valid():
                logger.info(f"Создано новое мероприятие из импорта {event_name}")
                serializer.save()
                self.created += 1
            else:
                self.errors.append(
                    {"row_number": row_number, "errors": serializer.errors}
                )

        self.workbook.close()
        logger.info(
            f"Импорт мероприятий завершен. Создано {self.created}, ошибок: {len(self.errors)}"
        )
        return {"created": self.created, "errors": self.errors}
